#!/usr/bin/env python3 -u
"""华为应用市场评论采集 - 使用适配器类

基于更新后的 HuaweiAppStoreAdapter，支持:
- 条数限制
- 天数筛选
- Excel 导出
- 进度显示

用法:
    # 测试模式 (10条)
    python scripts/adapter_scrape.py test

    # 采集指定数量
    python scripts/adapter_scrape.py 100

    # 采集指定天数内的评论
    python scripts/adapter_scrape.py --days 2

    # 组合使用
    python scripts/adapter_scrape.py 200 --days 2
"""
import sys
import argparse
from pathlib import Path
from datetime import datetime

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.device.connector import DeviceConnector
from src.adapters.huawei import HuaweiAppStoreAdapter, ScrapeConfig
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_excel(comments, filename):
    """创建 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "评论数据"

    # 表头
    headers = ["序号", "用户ID", "评论内容", "原始时间", "时间类型", "几天前", "采集时间"]
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 数据行
    collection_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for idx, comment in enumerate(comments, 1):
        days_ago_str = str(comment.days_ago) if comment.days_ago is not None else "未知"
        ws.append([
            idx,
            comment.user_id,
            comment.content,
            comment.time_str,
            comment.time_type,
            days_ago_str,
            collection_time
        ])

        # 应用边框
        for col_num in range(1, 8):
            ws.cell(row=idx + 1, column=col_num).border = thin_border

        # 根据天数设置颜色
        if comment.days_ago == 0:
            fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            for col_num in range(1, 8):
                ws.cell(row=idx + 1, column=col_num).fill = fill
        elif comment.days_ago == 1:
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for col_num in range(1, 8):
                ws.cell(row=idx + 1, column=col_num).fill = fill

    # 调整列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18

    # 自动换行
    for row in ws.iter_rows(min_row=2, max_row=len(comments) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 设置行高
    for row in range(2, len(comments) + 2):
        ws.row_dimensions[row].height = 60

    wb.save(filename)
    print(f"\n✓ Excel 已保存: {filename}")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="华为应用市场评论采集工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s test              # 测试模式，采集10条
  %(prog)s 100               # 采集100条
  %(prog)s --days 2          # 采集2天内的所有评论
  %(prog)s 200 --days 2      # 采集2天内最多200条
        """
    )

    parser.add_argument(
        "count",
        nargs="?",
        default="100",
        help="采集数量 (或 'test' 进行测试)"
    )

    parser.add_argument(
        "--days", "-d",
        type=int,
        default=None,
        help="只采集N天内的评论"
    )

    parser.add_argument(
        "--game", "-g",
        default="王者荣耀",
        help="游戏名称"
    )

    parser.add_argument(
        "--package", "-p",
        default="com.tencent.tmgp.sgame",
        help="应用包名"
    )

    args = parser.parse_args()

    # 解析数量参数
    if args.count.lower() == "test":
        max_count = 10
        test_mode = True
    else:
        try:
            max_count = int(args.count)
            test_mode = False
        except ValueError:
            print(f"错误: 无效的数量 '{args.count}'")
            return 1

    # 输出目录
    output_dir = Path("game-comment-scraper/data")
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    if test_mode:
        print("测试模式: 采集 10 条评论")
    else:
        print(f"采集模式: 最多 {max_count} 条评论")
        if args.days:
            print(f"时间限制: {args.days} 天内")
    print("=" * 60)
    print(f"游戏: {args.game}")
    print(f"包名: {args.package}")

    # 连接设备
    print("\n[1/3] 连接设备...")
    try:
        device = DeviceConnector()
        device.connect_usb()
        print(f"  ✓ 设备: {device.device.device_info['model']}")
    except Exception as e:
        print(f"  ✗ 设备连接失败: {e}")
        return 1

    # 创建适配器
    print("\n[2/3] 初始化适配器...")
    adapter = HuaweiAppStoreAdapter(device)

    # 配置采集参数
    config = ScrapeConfig(
        max_count=max_count,
        max_days=args.days,
        max_scrolls=200 if not test_mode else 30,
        progress_interval=5
    )

    # 开始采集
    print("\n[3/3] 开始采集...")
    print("-" * 60)

    try:
        comments = adapter.scrape_game_comments(
            game_name=args.game,
            package_name=args.package,
            config=config
        )

        print("-" * 60)

        # 统计信息
        days_stats = {}
        for c in comments:
            if c.days_ago is not None:
                days_stats[c.days_ago] = days_stats.get(c.days_ago, 0) + 1

        print(f"\n采集完成: {len(comments)} 条评论")

        if days_stats:
            print(f"\n天数分布:")
            for days in sorted(days_stats.keys()):
                label = "今天" if days == 0 else ("昨天" if days == 1 else f"{days}天前")
                print(f"  {label}: {days_stats[days]} 条")

            if args.days:
                within = sum(count for days, count in days_stats.items() if days <= args.days)
                print(f"\n  【{args.days}天内小计: {within} 条】")

        # 导出 Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if test_mode:
            filename = output_dir / f"test_adapter_{timestamp}.xlsx"
        else:
            days_suffix = f"_{args.days}days" if args.days else ""
            filename = output_dir / f"huawei_{args.game}_{len(comments)}comments{days_suffix}_{timestamp}.xlsx"

        create_excel(comments, str(filename))

        # 返回首页
        adapter.go_home()

        print("\n" + "=" * 60)
        print("任务完成!")
        print("=" * 60)
        return 0

    except Exception as e:
        print(f"\n✗ 采集失败: {e}")
        import traceback
        traceback.print_exc()
        adapter.go_home()
        return 1


if __name__ == "__main__":
    sys.exit(main())
