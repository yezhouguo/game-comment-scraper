#!/usr/bin/env python3 -u
"""华为应用市场 - 爬取评论并导出到Excel

说明: 华为应用市场时间格式限制
- 今天: HH:MM 格式 (如 08:02)
- 昨天: "昨天"
- 更早: MM-DD 或其他格式

由于 HH:MM 无法精确确定日期，本脚本采用以下策略:
1. 采集固定数量的评论，确保覆盖最近几天
2. 在Excel中标注时间类型供用户参考

用法:
    # 测试模式：爬取10条评论测试
    python scripts/scrape_to_excel.py test

    # 生产模式：爬取指定数量的评论
    python scripts/scrape_to_excel.py <数量>
    例如: python scripts/scrape_to_excel.py 500
"""
import uiautomator2 as u2
import time
import subprocess
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

# 配置
GAME_NAME = "王者荣耀"
PACKAGE_NAME = "com.tencent.tmgp.sgame"
OUTPUT_DIR = Path("game-comment-scraper/data")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 解析命令行参数
if len(sys.argv) < 2:
    print("用法:")
    print("  python scripts/scrape_to_excel.py test      # 测试模式：10条评论")
    print("  python scripts/scrape_to_excel.py <数量>   # 爬取指定数量的评论")
    print("  例如: python scripts/scrape_to_excel.py 500  # 爬取500条")
    sys.exit(1)

arg = sys.argv[1]

if arg.lower() == "test":
    TEST_MODE = True
    MAX_COMMENTS = 10
    TARGET_SCROLLS = 20
    print("=" * 60)
    print("测试模式：将爬取 10 条评论进行测试")
    print("=" * 60)
else:
    TEST_MODE = False
    try:
        MAX_COMMENTS = int(arg)
        # 根据数量估算滚动次数（平均每次滚动约获取3-5条新评论）
        TARGET_SCROLLS = max(100, MAX_COMMENTS // 3 * 2)
        print("=" * 60)
        print(f"生产模式：将爬取约 {MAX_COMMENTS} 条评论")
        print(f"预估滚动次数: {TARGET_SCROLLS}")
        print("=" * 60)
    except ValueError:
        print(f"错误：无效的数量参数 '{arg}'")
        sys.exit(1)


def estimate_time_type(time_str: str) -> str:
    """判断时间类型"""
    if not time_str:
        return "未知"
    elif "昨天" in time_str:
        return "昨天"
    elif "前" in time_str:
        return "相对时间"
    elif ":" in time_str and "/" not in time_str:
        return "今天(HH:MM)"
    elif "/" in time_str:
        return "完整日期"
    elif "-" in time_str and len(time_str) < 10:
        return "日期(MM-DD)"
    else:
        return "其他"


def parse_time_to_days_ago(time_str: str) -> int | None:
    """解析时间字符串，返回几天前

    返回值:
        None: 无法解析（假设是最近的）
        0: 今天
        1: 昨天
        n: n天前
    """
    if not time_str:
        return None

    today = datetime.now().date()

    try:
        # "昨天"
        if "昨天" in time_str:
            return 1

        # "2026/3/9" 格式
        if "/" in time_str and "年" not in time_str:
            parts = time_str.split("/")
            if len(parts) == 3:
                year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                comment_date = datetime(year, month, day).date()
                delta = (today - comment_date).days
                return max(0, delta)

        # "YYYY-MM-DD" 格式
        if "-" in time_str and len(time_str) >= 10:
            comment_date = datetime.strptime(time_str, "%Y-%m-%d").date()
            delta = (today - comment_date).days
            return max(0, delta)

        # "MM-DD" 格式（假设是今年）
        if "-" in time_str and len(time_str) < 10:
            parts = time_str.split("-")
            if len(parts) == 2:
                month, day = int(parts[0]), int(parts[1])
                comment_date = datetime(today.year, month, day).date()
                # 如果日期在未来，说明是去年的
                if comment_date > datetime.now().date():
                    comment_date = datetime(today.year - 1, month, day).date()
                delta = (today - comment_date).days
                return max(0, delta)

        # "HH:MM" 格式 - 假设是今天
        if ":" in time_str:
            return 0

    except Exception:
        pass

    return None


def create_excel(comments: list, filename: str):
    """创建Excel文件"""

    wb = Workbook()
    ws = wb.active
    ws.title = "评论数据"

    # 表头
    headers = ["序号", "用户ID", "评论内容", "原始时间", "时间类型", "采集时间"]
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 数据行
    collection_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for idx, comment in enumerate(comments, 1):
        time_type = estimate_time_type(comment['time_str'])
        ws.append([
            idx,
            comment['user_id'],
            comment['content'],
            comment['time_str'],
            time_type,
            collection_time
        ])

        # 应用边框
        for col_num in range(1, 7):
            ws.cell(row=idx + 1, column=col_num).border = thin_border

        # 根据时间类型设置颜色
        if "昨天" in time_type:
            for col_num in range(1, 7):
                ws.cell(row=idx + 1, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif "今天" in time_type:
            for col_num in range(1, 7):
                ws.cell(row=idx + 1, column=col_num).fill = PatternFill(
                    start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
                )

    # 调整列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # 自动换行
    for row in ws.iter_rows(min_row=2, max_row=len(comments) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 设置行高
    for row in range(2, len(comments) + 2):
        ws.row_dimensions[row].height = 60

    wb.save(filename)
    print(f"\n✓ Excel 文件已保存: {filename}")
    print(f"  - 绿色行: 今天的评论")
    print(f"  - 黄色行: 昨天的评论")


# ==================== 主程序 ====================

print(f"\n目标游戏: {GAME_NAME}")
print(f"包名: {PACKAGE_NAME}")

# 连接设备
print("\n[1/4] 连接设备...")
d = u2.connect_usb()
print(f"    ✓ 设备连接成功")

# 导航到评论页面
print("\n[2/4] 导航到评论页面...")
d.press("home")
time.sleep(1)

url = f"market://details?id={PACKAGE_NAME}"
cmd = ["adb", "shell", "am", "start", "-a", "android.intent.action.VIEW", "-d", url, "com.huawei.appmarket"]
subprocess.run(cmd, capture_output=True)
time.sleep(3)

# 点击"介绍"
intro = d(resourceId="com.huawei.appmarket:id/other_appinfos")
b = intro.info.get('bounds', {})
x = b.get('left', 318) + (b.get('right', 483) - b.get('left', 318)) * 0.3
y = (b.get('top', 315) + b.get('bottom', 343)) / 2
d.click(int(x), int(y))
time.sleep(2)

# 点击"评论"
d(textContains="评论").click()
time.sleep(2)
print("    ✓ 已进入评论页面")

# 采集评论
print("\n[3/4] 开始采集评论...")
print("-" * 60)

all_comments = []
seen = set()
scroll_count = 0
consecutive_no_new = 0

start_time = time.time()

# 统计时间类型
time_stats = {}
days_stats = {}  # 按天数统计

while scroll_count < TARGET_SCROLLS:
    # 获取所有评论元素
    users = d(resourceId="com.huawei.appmarket:id/detail_comment_user_textview")
    contents = d(resourceId="com.huawei.appmarket:id/detail_comment_content_textview")
    times = d(resourceId="com.huawei.appmarket:id/detail_comment_time_textview")

    min_count = min(len(users), len(contents), len(times))

    new_added = 0
    batch_days_ago = []  # 批次中评论的天数

    for i in range(min_count):
        uid = users[i].info.get('text', '')
        cont = contents[i].info.get('text', '')
        time_str = times[i].info.get('text', '')

        # 去重（基于内容）
        if not cont or cont in seen:
            continue

        # 解析天数
        days_ago = parse_time_to_days_ago(time_str)
        if days_ago is not None:
            batch_days_ago.append(days_ago)

        # 统计时间类型
        time_type = estimate_time_type(time_str)
        time_stats[time_type] = time_stats.get(time_type, 0) + 1

        # 统计天数
        if days_ago is not None:
            days_stats[days_ago] = days_stats.get(days_ago, 0) + 1

        seen.add(cont)
        all_comments.append({
            'user_id': uid,
            'content': cont,
            'time_str': time_str,
            'days_ago': days_ago
        })
        new_added += 1

    # 进度报告（每5次滚动报告一次）
    if scroll_count % 5 == 0:
        type_info = ", ".join([f"{k}:{v}" for k, v in time_stats.items()])
        # 显示最旧的评论天数
        max_days = max(batch_days_ago) if batch_days_ago else 0
        print(f"    滚动 {scroll_count}: 已采集 {len(all_comments)} 条 | 最旧: {max_days}天前 | ({type_info})", flush=True)

        # 如果最旧的评论已经超过2天，且已经采集了足够数据，可以考虑停止
        if max_days > 2 and len(all_comments) >= 100:
            print(f"    已采集到超过2天的评论，当前共 {len(all_comments)} 条")

    # 测试模式：达到目标数量后停止
    if TEST_MODE and len(all_comments) >= MAX_COMMENTS:
        break

    # 生产模式：检查是否达到目标数量
    if not TEST_MODE and len(all_comments) >= MAX_COMMENTS:
        print(f"    已达到目标数量 {MAX_COMMENTS} 条，停止采集")
        break

    # 连续多次没有新评论，可能到底了
    if new_added == 0 and scroll_count > 10:
        consecutive_no_new += 1
        if consecutive_no_new >= 5:
            print("    没有新评论，已到达底部")
            break
    else:
        consecutive_no_new = 0

    # 滚动加载更多
    d.swipe(0.5, 0.75, 0.5, 0.25, duration=0.3)
    time.sleep(1.5)
    scroll_count += 1

elapsed = time.time() - start_time

print("-" * 60)
print(f"\n采集完成!")
print(f"    总评论数: {len(all_comments)} 条")
print(f"    总滚动次数: {scroll_count}")
print(f"    总耗时: {elapsed:.1f} 秒")
if len(all_comments) > 0:
    print(f"    平均速度: {elapsed/len(all_comments):.2f} 秒/条")

# 显示时间类型统计
print(f"\n时间类型分布:")
for time_type, count in sorted(time_stats.items(), key=lambda x: -x[1]):
    pct = count / len(all_comments) * 100 if all_comments else 0
    print(f"    {time_type}: {count} 条 ({pct:.1f}%)")

# 显示按天数统计（2天内）
print(f"\n按天数统计:")
within_2_days = 0
for days in sorted(days_stats.keys()):
    count = days_stats[days]
    if days == 0:
        label = "今天"
    elif days == 1:
        label = "昨天"
    else:
        label = f"{days}天前"
    print(f"    {label}: {count} 条")
    if days <= 2:
        within_2_days += count

print(f"\n    【2天内小计: {within_2_days} 条】")
print(f"    【3天及以上: {len(all_comments) - within_2_days} 条】")

# 生成Excel
print("\n[4/4] 生成Excel文件...")
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
if TEST_MODE:
    filename = OUTPUT_DIR / f"test_comments_{timestamp}.xlsx"
else:
    filename = OUTPUT_DIR / f"huawei_{GAME_NAME}_{len(all_comments)}comments_{timestamp}.xlsx"

create_excel(all_comments, str(filename))

# 返回首页
d.press("home")
time.sleep(1)

print("\n" + "=" * 60)
print("任务完成!")
print(f"  共采集 {len(all_comments)} 条评论")
print(f"  Excel 文件: {filename}")
print("=" * 60)
